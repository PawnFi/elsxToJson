from copy import copy
from posixpath import split
import shutil
from numpy import append
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell
import json_tools
import json
import io
import os
import msvcrt
def excel_to_jsons(excel_file,file_name):
    book = openpyxl.load_workbook(excel_file)
    sheet = book["Sheet1"]
    max_row = sheet.max_row
    max_column = sheet.max_column
    TitleList=[]
    LanList = []
    version_column = 0
    module_column = 0
    component_column = 0
    key_column = 0
    zh_column = 0 #--
    en_column = 0 #--
    fix_column = 0
    
    for i in range(1,max_column):
        temp = sheet.cell(1,i).value
        if temp!=None:
            TitleList.append({temp:i})
            if temp == 'version':version_column = i
            elif temp == 'module':module_column = i
            elif temp == '二级':component_column = i
            elif temp == '三级':key_column = i
            # elif temp == 'zh':zh_coloum = i
            # elif temp == 'en':en_coloum = i  
            elif temp == '修訂版':fix_column = i   
            elif i>max(1,key_column): LanList.append({temp:i})
    max_end_row = getMaxEndRow(sheet,module_column)
    print("生成{}种语言({}):".format(len(LanList),file_name))#--
    for i in range(len(LanList)):
        lan_name = list(LanList[i].keys())[0]
        lan_column = list(LanList[i].values())[0]
        print(lan_name.upper()+':')#--
        excel_to_json(sheet,lan_name,max_end_row,module_column,component_column,key_column,lan_column,file_name)
    book.close()
            
def getMaxEndRow(sheet,dependcolumn):
    max_row = sheet.max_row
    max_end_row = 0
    flag = ''
    for i in range(2,max_row):
        temp = sheet.cell(i,dependcolumn)
        if isinstance(temp,MergedCell)==False:
            # print(temp.value,flag)
            if temp.value != flag:
                flag = temp.value
                max_end_row = i 
    # print(max_end_row)  
    return max_end_row         


def excel_to_json(sheet,lan_name,max_end_row,module_column,component_coloum,key_column,lan_column,file_name):
    # result = {}
    l1 = getSortList(sheet,2,max_end_row,module_column) 
    for index in range(len(l1)):
        subRes = {}
        l2 = getSortList(sheet,l1[index]["start"],l1[index]["end"],component_coloum)
        for i in range(len(l2)):
            value = getLanDetail(sheet,l2[i]["start"],l2[i]["end"],key_column,lan_column)
            key = l2[i]['head']
            if key != None:
                subRes[key]=value
                
        pathName='./output/{}/{}/'.format(file_name,lan_name)
        json_file_name =pathName+ '{}.json'.format(l1[index]['head'])
        old_pathName = pathName+'old/'
        old_json_file_name = old_pathName+ 'old{}.json'.format(l1[index]['head'])
        change_pathName = pathName+'old/change/'
        change_json_file_name = change_pathName+ 'change{}.json'.format(l1[index]['head'])
        if not os.path.exists(json_file_name): # 不存在这个文件 
            if not os.path.exists(pathName):  # 不存在这个文件夹 
                os.makedirs(pathName)  # 如果没有这个文件夹，那就创建一个
            save_json_file(subRes, json_file_name)# 创建新文件
        else:                                  # 存在这个文件
            if not os.path.exists(old_pathName):  # 不存在这个文件夹 
                os.makedirs(old_pathName)  # 如果没有这个文件夹，那就创建一个
            if os.path.exists(old_json_file_name):os.remove(old_json_file_name)
            os.rename(json_file_name,old_json_file_name) #创建旧文件(已有文件改名并移动)
            
            save_json_file(subRes, json_file_name)# 创建新文件
            
            if not os.path.exists(change_pathName):  # 不存在这个文件夹 
                os.makedirs(change_pathName)  # 如果没有这个文件夹，那就创建一个
            getJsonDiff(old_json_file_name,json_file_name,change_json_file_name) #生成对比文件    

def getSortList(sheet,start=2,stop=389,column=3):
    tempList = []
    heads = []
    for row in range(start,stop+1):
        temp=sheet.cell(row,column)
        if(isinstance(temp,MergedCell)==False):
            tempList.append(row)
        if(temp.value!= None):
            heads.append(temp.value)
    # print(len(heads)*2,len(tempList))
    if((len(heads)*2 != len(tempList)) and column<=3):print('在第{}列发现一个格式错误，生成json可能存在问题，请修复！'.format(column))
    tempList = tempList[slice(0,len(heads)*2)]
    sortList = []
    tempItem = {}
    for index in range(len(tempList)):
        if index % 2 == 0:
            tempItem['head'] = heads[int(index/2)]
            tempItem['start'] = tempList[index]
        else:
            tempItem['end'] = tempList[index]
            sortList.append(copy(tempItem))
    # print(sortList)
    return sortList

def getLanDetail(sheet,start,end,key_column=5,lan_column=7):
    result = {}
    for i in range(start,end):
        key = transNUllWithTips(sheet.cell(i,key_column).value,i,key_column,sheet.cell(i,lan_column))
        if key != None:
            result[key] = transNull(sheet.cell(i,lan_column).value)
    return result
def transNull(nul):
    if nul == None:
        return ''
    else:
        return nul
def transNUllWithTips(nul,row,column,cell):
    if nul == None and cell.value != None:
        print('在{}行{}列缺失key，请检查。'.format(row,column))
    return transNull(nul)

def save_json_file(jd,json_file_name):
    
    file = io.open(json_file_name,'w',encoding='utf-8')
    txt = json.dumps(jd, indent=2, ensure_ascii=False)
    file.write(txt)
    file.close()

def getJsonDiff(oldpath,newpath,changepath):
    oldJson={}
    newJson={}
    with open(oldpath,'r',encoding='utf-8') as f:
        oldJson = json.load(f)
    with open(newpath,'r',encoding='utf-8') as f:
        newJson = json.load(f)
    diffData = json_tools.diff(oldJson,newJson)
    # print(diffData)
    for i in range(len(diffData)):
        item = diffData[i]
        key = list(item.keys())[0]
        # print(diffData[i],key)
        if key == 'add':
            if not isinstance(item['value'],str):
                item['value']={'新增对象键与内容':item['value']}
            else: item['value'] = item['value']+ ' //新增键与内容++++++++++++++++++++++++++++'
        elif key == 'replace':
            if item['prev'] == '':
                if not isinstance(item['value'],str):
                    item['value']={'新增对象内容':item['value']}
                else: item['value'] = item['value']+ ' //新增内容++++++++++++++++++++++++++++'
            elif item['value'] == '':
                if not (isinstance(item['value'],str) or isinstance(item['prev'],str)):
                    item['value']={'删除对象内容':item['prev']}
                else: item['value'] = item['value']+ ' //删除内容-------------------------prev: '+item['prev']
            else:
                if not (isinstance(item['value'],str) or isinstance(item['prev'],str)):
                    item['value']={'更改对象内容':item['prev']}
                else: item['value'] = item['value']+ ' //更改============================prev: '+item['prev']
        elif key == 'remove':
            item['add'] = item['remove']
            del(item['remove'])
            if not isinstance(item['prev'],str):
                item['value']={'删除对象键与内容':item['prev']}
            else: item['value']=' //删除键与内容-------------------------prev: '+item['prev']
            del(item['prev'])
        diffData[i] = copy(item)
        # print(item)
    # print(diffData)
    if diffData == False:print(newpath+"发生更改，请打开"+changepath+'检查更改')
    
    changeJson = json_tools.patch(oldJson,diffData)
    save_json_file(changeJson,changepath)
    
def ScanFile(directory, prefix=None, postfix=None):
    file_list = []
    temp_list = []
    for root, sub_dirs, files in os.walk(directory):
        for special_file in files:
            # 如果指定前缀或者后缀
            if postfix or prefix:
                # 同时指定前缀和后缀
                if postfix and prefix:
                    if special_file.endswith(postfix) and special_file.startswith(prefix):
                        file_list.append(os.path.join(root, special_file))
                        temp_list.append(os.path.join(special_file))
                        continue

                # 只指定后缀
                elif postfix:
                    if special_file.endswith(postfix):
                        file_list.append(os.path.join(root, special_file))
                        temp_list.append(os.path.join(special_file))
                        continue

                # 只指定前缀
                elif prefix:
                    if special_file.startswith(prefix):
                        file_list.append(os.path.join(root, special_file))
                        # temp_list.append(os.path.join(special_file))
                        continue

            # 前缀后缀均未指定
            else:
                file_list.append(os.path.join(root, special_file))
                continue
	# 打印出扫描到的文件路径
    # print(file_list,temp_list)
    return [file_list,temp_list]
            
def outPut():
    prefix = input("请输入筛选前缀:\n")
    temp = ScanFile("./input",prefix,".xlsx")
    for index,item in enumerate(temp[0]):
        file_name = temp[1][index].split('.')[0]
        # print(item,file_name)
        excel_to_jsons(item,file_name) 
    print("请按任意键退出~")
    ord(msvcrt.getch())

if '__main__'==__name__:
    # excel_to_jsons(u'./input/testPro.xlsx')
    outPut()
    